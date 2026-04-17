import streamlit as st
import anthropic
import openpyxl
import re
import time
import io
from pathlib import Path

# ── 頁面設定 ──────────────────────────────────────────
st.set_page_config(
    page_title="中文 → 日文 Excel 翻譯器",
    page_icon="🗾",
    layout="centered"
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Noto+Serif+JP:wght@300;400;600&family=Noto+Serif+TC:wght@300;400;600&display=swap');
html, body, [class*="css"] { font-family: 'Noto Serif TC', 'Noto Serif JP', serif; }
.title-jp { font-family: 'Noto Serif JP', serif; font-size: 13px; letter-spacing: 6px; color: #7a6f60; margin-bottom: 8px; text-align: center; }
.title-main { font-size: 2.2rem; font-weight: 600; color: #1a1a2e; letter-spacing: 2px; text-align: center; }
.title-main span { color: #c0392b; }
.subtitle { font-size: 14px; color: #7a6f60; letter-spacing: 1px; margin-top: 6px; text-align: center; }
.stat-box { background: white; border: 1px solid #d4c9b0; border-radius: 4px; padding: 1rem; text-align: center; }
.stat-num { font-size: 2rem; font-weight: 600; color: #1a1a2e; }
.stat-lbl { font-size: 11px; letter-spacing: 2px; color: #7a6f60; text-transform: uppercase; }
.stButton > button { background-color: #1a1a2e !important; color: #f5f0e8 !important; font-family: 'Noto Serif TC', serif !important; letter-spacing: 3px !important; border: none !important; width: 100% !important; font-size: 15px !important; }
.stButton > button:hover { background-color: #c0392b !important; }
</style>
""", unsafe_allow_html=True)

# ── 標題 ──────────────────────────────────────────────
st.markdown('<div class="title-jp">中文 → 日本語 翻訳ツール</div>', unsafe_allow_html=True)
st.markdown('<div class="title-main">Excel <span>中→日</span> 翻譯器</div>', unsafe_allow_html=True)
st.markdown('<div class="subtitle">上傳 Excel，以日本會計實務用語自動翻譯成日文・格式完整保留</div>', unsafe_allow_html=True)
st.divider()

# ── 設定 ──────────────────────────────────────────────
BATCH_SIZE = 20
MODEL = "claude-sonnet-4-6"
MODEL_PRICING = {
    "claude-sonnet-4-6":         {"input": 3.0,  "output": 15.0},
    "claude-haiku-4-5-20251001": {"input": 0.8,  "output": 4.0},
    "claude-opus-4-6":           {"input": 15.0, "output": 75.0},
}

TERM_GLOSSARY = """
【強制術語對照表－請嚴格遵守以下對應，不得使用其他譯法】
貸借対照表　← 資產負債表
損益計算書　← 損益表
キャッシュ・フロー計算書　← 現金流量表
株主資本等変動計算書　← 股東權益變動表
連結財務諸表　← 合併財務報表
個別財務諸表　← 個別財務報表
財務諸表注記　← 財務報表附註
試算表（T/B）　← 試算表
売掛金　← 應收帳款
受取手形　← 應收票據
棚卸資産　← 存貨
前払金　← 預付款項
その他流動資産　← 其他流動資產
有形固定資産　← 不動產廠房設備
使用権資産　← 使用權資產
無形資産　← 無形資產
繰延税金資産　← 遞延所得稅資產
投資不動産　← 投資性不動產
買掛金　← 應付帳款
支払手形　← 應付票據
短期借入金　← 短期借款
長期借入金　← 長期借款
未払費用　← 應付費用
前受金　← 預收款項
リース負債　← 租賃負債
繰延税金負債　← 遞延所得稅負債
その他流動負債　← 其他流動負債
売上高／営業収益　← 營業收入
売上原価　← 銷售成本
売上総利益　← 毛利
営業費用　← 營業費用
販売費　← 銷售費用
一般管理費　← 管理費用
研究開発費　← 研究發展費用
営業利益　← 營業利益
営業外収益　← 業外收入
営業外費用　← 業外支出
税引前当期純利益　← 稅前淨利
法人税等　← 所得稅費用
当期純利益　← 本期淨利
減価償却費　← 折舊
償却費　← 攤銷
従業員福利費用　← 員工福利費用
仕訳　← 分錄
調整仕訳　← 調整分錄
期首残高　← 期初餘額
期末残高　← 期末餘額
帳簿計上額　← 帳列數
帳簿価額　← 帳面金額
相殺消去前帳簿計上額　← 抵銷前帳面金額
小計　← 小計
合計　← 合計
借方　← 借方
貸方　← 貸方
転記　← 過帳
取り消し／逆仕訳　← 沖銷／迴轉
関連当事者　← 關聯方／關係人
関連当事者取引　← 關係人交易
親会社　← 母公司
子会社　← 子公司
連結子会社　← 連結子公司
グループ　← 集團
内部取引　← 內部交易
消去　← 消去
連結プロセス　← 連結作業
持株比率　← 持股比例
アップロード　← 上傳
ダウンロード　← 下載
データを保存　← 儲存資料
名前を付けて保存　← 另存新檔
フォルダ　← 資料夾
レポート　← 報表
パッケージ　← 套件
確認済　← 確認完成
入力ステータス　← 輸入狀態／填寫狀態
目次に戻る　← 回到目錄
青色グリッドエリア　← 藍色網格區域
原通貨　← 原始貨幣
機能通貨　← 功能性貨幣
担当責任者　← 負責人／担当者
重要性　← 重大性
内部統制　← 內部控制
監査　← 查核
レビュー　← 複核
陳述書　← 聲明書
限定意見　← 保留意見
会計方針　← 會計政策
会計上の見積り　← 會計估計
重要な会計項目　← 重大會計項目
「確認済」を選択してデータを保存　← 選擇確認完成並儲存資料
"""

# ── 工具函數 ──────────────────────────────────────────
def convert_roc_date(text: str) -> str:
    def replace_year(m):
        roc_year = int(m.group(1))
        if roc_year < 200:
            return str(roc_year + 1911) + "年"
        return m.group(0)
    return re.sub(r"(?<!\d)(\d{2,3})年", replace_year, text)

def has_chinese(text: str) -> bool:
    return bool(re.search(r'[\u4e00-\u9fff\u3400-\u4dbf]', str(text)))

def collect_cells(wb):
    cells = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    val = cell.value
                    if has_chinese(val) or re.search(r'(?<!\d)\d{2,3}年', val):
                        cells.append((sheet_name, cell.row, cell.column, val))
    return cells

def estimate_cost(cells, model):
    total_chars = sum(len(c[3]) for c in cells)
    total_batches = (len(cells) + BATCH_SIZE - 1) // BATCH_SIZE
    input_tokens  = int(total_chars * 1.5) + total_batches * 800
    output_tokens = int(total_chars * 1.5)
    pricing = MODEL_PRICING.get(model, {"input": 3.0, "output": 15.0})
    total_cost = (input_tokens / 1_000_000 * pricing["input"] +
                  output_tokens / 1_000_000 * pricing["output"])
    return total_chars, input_tokens, output_tokens, total_cost

def translate_batch(texts, client, model):
    joined = "\n§§§\n".join(texts)
    message = client.messages.create(
        model=model,
        max_tokens=4000,
        messages=[{
            "role": "user",
            "content": f"""あなたは日本の公認会計士（CPA）資格を持つ、財務・会計専門の翻訳者です。
以下の中国語テキストを、日本の会計実務で使用される正式な用語・表現に翻訳してください。

{TERM_GLOSSARY}

【翻訳ルール】
- 上記の術語対照表に該当する用語は必ずその日本語訳を使用すること
- 会計・財務用語は日本GAAP・税務実務の正式用語を優先する
- 数字・記号・英文はそのまま保持する
- 単純な項目名・ラベルは簡潔に訳す（余分な説明を加えない）
- 文脈から会計用語でないと判断できる場合は自然な日本語にする

【重要：出力形式】
- 各テキストは「§§§」で区切られている。翻訳後も必ず同じ「§§§」で区切って出力すること
- 番号・説明・原文は出力しない
- 区切り記号「§§§」を追加・削除・変更しないこと（入力と同じ {len(texts)} ブロックを出力）

{joined}"""
        }]
    )
    result = message.content[0].text.strip()
    blocks = [b.strip() for b in result.split("§§§")]
    while len(blocks) < len(texts):
        blocks.append(texts[len(blocks)])
    return blocks[:len(texts)]

# ── 介面 ──────────────────────────────────────────────
col_key, col_model = st.columns([3, 2])
with col_key:
    api_key = st.text_input("🔑 Claude API Key", type="password", placeholder="sk-ant-api03-...")
with col_model:
    selected_model = st.selectbox("模型", list(MODEL_PRICING.keys()),
        format_func=lambda x: {"claude-sonnet-4-6": "Sonnet（推薦）",
                                "claude-haiku-4-5-20251001": "Haiku（省費用）",
                                "claude-opus-4-6": "Opus（最強）"}.get(x, x))

translate_sheet_names = st.checkbox("同時翻譯工作表名稱", value=True)
uploaded_file = st.file_uploader("📊 上傳 Excel 檔案", type=["xlsx", "xls"])

st.divider()

if uploaded_file:
    try:
        wb_preview = openpyxl.load_workbook(io.BytesIO(uploaded_file.read()))
        uploaded_file.seek(0)
        cells_preview = collect_cells(wb_preview)
        total_preview = len(cells_preview)
        total_chars, input_tok, output_tok, est_cost = estimate_cost(cells_preview, selected_model)
        total_batches = (total_preview + BATCH_SIZE - 1) // BATCH_SIZE

        st.markdown("**📊 翻譯預估**")
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("工作表", f"{len(wb_preview.sheetnames)} 個")
        c2.metric("含中文儲存格", f"{total_preview} 格")
        with c3:
            st.caption("預估費用")
            st.write(f"**USD ${est_cost:.4f}**")
        with c4:
            st.caption("約新台幣")
            st.write(f"**TWD ${est_cost * 32:.1f}**")
        st.caption(f"預估 Token：輸入 {input_tok:,} / 輸出 {output_tok:,}・預估批次：{total_batches} 批")
        st.divider()
    except Exception as e:
        st.error(f"無法讀取檔案：{e}")
        st.stop()

if st.button("🚀 開始翻譯"):
    if not api_key:
        st.error("請輸入 Claude API Key")
        st.stop()
    if not api_key.startswith("sk-ant"):
        st.error("API Key 格式不正確")
        st.stop()
    if not uploaded_file:
        st.error("請上傳 Excel 檔案")
        st.stop()

    try:
        client = anthropic.Anthropic(api_key=api_key)
        wb = openpyxl.load_workbook(io.BytesIO(uploaded_file.read()))
    except Exception as e:
        st.error(f"讀取失敗：{e}")
        st.stop()

    cells = collect_cells(wb)
    total = len(cells)
    if total == 0:
        st.warning("找不到任何含中文的儲存格！")
        st.stop()

    progress_bar = st.progress(0)
    status_text  = st.empty()
    log_area     = st.empty()

    done = 0
    errors = 0
    total_batches = (total + BATCH_SIZE - 1) // BATCH_SIZE
    log_lines = []

    for i in range(0, total, BATCH_SIZE):
        batch     = cells[i:i + BATCH_SIZE]
        batch_num = i // BATCH_SIZE + 1

        date_only        = [(j, c) for j, c in enumerate(batch) if not has_chinese(c[3])]
        translate_needed = [(j, c) for j, c in enumerate(batch) if has_chinese(c[3])]

        for _, (sheet_name, row, col, val) in date_only:
            wb[sheet_name].cell(row=row, column=col).value = convert_roc_date(val)

        if not translate_needed:
            done += len(batch)
            progress_bar.progress(done / total)
            status_text.text(f"批次 {batch_num}/{total_batches}・{done}/{total} 完成（日期轉換）")
            continue

        texts = [c[3] for _, c in translate_needed]
        status_text.text(f"翻譯中… 批次 {batch_num}/{total_batches}・{done}/{total} 完成")

        for attempt in range(2):
            try:
                translations = translate_batch(texts, client, selected_model)
                for (_, (sheet_name, row, col, _)), translated in zip(translate_needed, translations):
                    wb[sheet_name].cell(row=row, column=col).value = convert_roc_date(translated)
                done += len(batch)
                log_lines.append(f"✅ 批次 {batch_num}/{total_batches} 完成（{len(batch)} 格）")
                break
            except Exception as e:
                if attempt == 0:
                    log_lines.append(f"⚠️ 批次 {batch_num} 錯誤，重試中… ({e})")
                    time.sleep(2)
                else:
                    log_lines.append(f"❌ 批次 {batch_num} 失敗，跳過")
                    errors += len(batch)
                    done   += len(batch)

        progress_bar.progress(done / total)
        log_area.code("\n".join(log_lines[-8:]))

        if i + BATCH_SIZE < total:
            time.sleep(0.3)

    # 翻譯工作表名稱
    if translate_sheet_names:
        chinese_names = [n for n in wb.sheetnames if has_chinese(n)]
        if chinese_names:
            status_text.text("翻譯工作表名稱…")
            try:
                translated_names = translate_batch(chinese_names, client, selected_model)
                for old, new in zip(chinese_names, translated_names):
                    wb[old].title = new
                    log_lines.append(f"📋 工作表：{old} → {new}")
                log_area.code("\n".join(log_lines[-8:]))
            except Exception as e:
                log_lines.append(f"⚠️ 工作表名稱翻譯失敗：{e}")

    status_text.text("✅ 翻譯完成！")
    progress_bar.progress(1.0)

    # 輸出檔案
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    output_name = Path(uploaded_file.name).stem + "_日本語.xlsx"

    st.success("🎉 翻譯完成！")
    ca, cb, cc = st.columns(3)
    ca.metric("已翻譯", f"{done - errors} 格")
    cb.metric("跳過", f"{errors} 格")
    cc.metric("工作表", f"{len(wb.sheetnames)} 個")

    st.download_button(
        label="⬇️ 下載翻譯後的 Excel",
        data=output,
        file_name=output_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

st.divider()
st.caption("由 Claude AI 驅動・日本 GAAP 会計実務用語対応・格式完整保留")
